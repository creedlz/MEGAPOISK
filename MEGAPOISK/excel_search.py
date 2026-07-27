import os
import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import load_workbook

import history

results = []

frame = None
folder_entry = None
search_entry = None
result_list = None
back_button = None


# выр (АХХХАХАХАХ в РФ такого нету АХАХАХАХАХА)
def choose_folder():
    folder = filedialog.askdirectory(
        title="Оберіть папку")
    if folder:
        folder_entry.delete(0, tk.END)
        folder_entry.insert(0, folder)

# сёртч
def search_files():
    results.clear()
    result_list.delete(0,tk.END)
    folder = folder_entry.get().strip()
    query = search_entry.get().strip().lower()
    if query:
     history.add(query)
    if not os.path.isdir(folder):
        messagebox.showerror(
            "Ошибка",
            "Папку не найдено")
        return
    if not query:
        messagebox.showwarning(
            "Внимание",
            "Введите текст для поиска")
        return



    for root, _, files in os.walk(folder):
        for fn in files:
            if fn.lower().endswith((".xlsx", ".xlsm")):
                path = os.path.join(root,fn)
                try:
                    wb = load_workbook(
                        path,
                        data_only=False)
                    for ws in wb.worksheets:
                        for row in ws.iter_rows():
                            for cell in row:
                                if cell.value is None:
                                    continue
                                text = str(cell.value)
                                if query in text.lower():
                                    results.append((path,))
                                    result_list.insert(
                                        tk.END,
                                        f"{fn} | Лист: {ws.title} | Строка: {cell.row} | Ячейка: {cell.coordinate} | {text[:80]}")
                except Exception as e:  # noqa: BLE001
                    print(e)
    if not results:
        result_list.insert(tk.END,"Ничего не найдено 😓")

# опен зе дор
def open_result(event=None):
    sel = result_list.curselection()
    if sel:
        os.startfile(results[sel[0]][0])



# скретч интерфейс
def init(parent):
    global frame
    global folder_entry
    global search_entry
    global result_list
    global back_button
    frame = tk.Frame(parent,bg="#202124")
    title = tk.Label(
        frame,
        text="📊 Поиск в Excel",
        font=("Segoe UI", 16, "bold"),
        fg="white",
        bg="#202124"
    )
    title.pack(pady=15)

    #папочка
    folder_frame = tk.Frame(frame,bg="#202124")
    folder_frame.pack(
        fill="x",
        padx=20,
        pady=5
    )
    folder_entry = tk.Entry(folder_frame)
    folder_entry.pack(
        side="left",
        fill="x",
        expand=True,
        padx=(0, 10)
    )
    tk.Button(
        folder_frame,
        text="📁 Выбрать",
        width=12,
        command=choose_folder).pack(side="right")

    #сертч туу
    search_frame = tk.Frame(frame,bg="#202124")
    search_frame.pack(
        fill="x",
        padx=20,
        pady=10
    )
    search_entry = tk.Entry(search_frame)
    search_entry.pack(
        side="left",
        fill="x",
        expand=True,
        padx=(0, 10)
    )
    tk.Button(
        search_frame,
        text="🔍 Найти",
        width=12,
        command=search_files).pack(side="right")
    
    # результот
    result_list = tk.Listbox(frame,font=("Consolas", 10))
    result_list.pack(fill="both", expand=True, padx=20, pady=10)
    
    result_list.bind("<Double-Button-1>",open_result)
    back_button = tk.Button(frame, text="⬅ Назад")

    tk.Button(
    frame,
    text="📂 Открыть файл",
    font=("Segoe UI", 10, "bold"),
    height=2,
    command=open_result).pack(fill="x", padx=20, pady=10)

    back_button = tk.Button(frame, text="⬅ Назад", width=12)
    back_button.pack(fill="x", padx=20, pady=(0, 15))

def show():
    frame.pack(fill="both", expand=True)
def hide():
    frame.pack_forget()
def set_back(command):
    global back_button  # noqa: PLW0602
    back_button.config(command=command)